#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build lại logic các đời gia phả từ file thô `Gia phả.xlsx`.

Cấu trúc bảng (Sheet1, dữ liệu bắt đầu từ dòng 3):
  A: Tên          (có thể chứa biến thể "Tên A/ Tên B")
  B: Ngày sinh    (dạng "d.m" - quy ước cũ coi là NGÀY GIỖ ÂM LỊCH)
  C: Vợ/Chồng     (tên phối ngẫu; có thể chứa biến thể)
  D: ngày của phối ngẫu
  E: Thế hệ thứ   (đời - dùng làm chuẩn, có ô trống)
  F/G: Giới tính Nam / Nữ (True/False)
  H: Cha / Mẹ     (cha mẹ của người ở cột A; chỉ điền ở "hàng chốt" mỗi nhóm)

Đầu ra:
  1) giapha-from-xlsx.json     - format import v2 của app (persons + relationships).
  2) giapha-from-xlsx-review.md - báo cáo đối soát logic các đời.

Run: python scripts/build-from-xlsx.py
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
import uuid
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "Gia phả.xlsx"
OUT_JSON = ROOT / "plans" / "260619-1933-digitize-genealogy" / "giapha-from-xlsx.json"
OUT_REPORT = ROOT / "plans" / "260619-1933-digitize-genealogy" / "giapha-from-xlsx-review.md"


def canonical(s: str) -> str:
    """Chuẩn hoá tên để so khớp: bỏ dấu, bỏ '()', bỏ ký tự lạ, hạ thường."""
    if not s:
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("(", "").replace(")", " ")
    s = re.sub(r"[^a-zA-Z0-9]+", " ", s)
    return " ".join(s.lower().split())


def split_variants(s: str | None) -> list[str]:
    """Tách biến thể tên theo dấu '/'. VD 'Hồ Việt Nhật/ Hồ Công Nhật'."""
    if not s:
        return []
    parts = re.split(r"[/]", s)
    return [p.strip(" /") for p in parts if p.strip(" /")]


def first_nonempty(*vals) -> str:
    for v in vals:
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


# ---------------------------------------------------------------------------
# Đọc bảng
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Sheet1"]

raw_rows = []
current_owner_row = None
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    gi = ws.cell(r, 2).value
    spouse = ws.cell(r, 3).value
    spouse_gi = ws.cell(r, 4).value
    gen = ws.cell(r, 5).value
    male = ws.cell(r, 6).value
    female = ws.cell(r, 7).value
    parent_cell = ws.cell(r, 8).value

    has_name = name is not None and str(name).strip()
    has_spouse = spouse is not None and str(spouse).strip()

    if has_name:
        try:
            m = bool(male)
            f = bool(female)
        except Exception:
            m = f = False
        gender = "male" if (m and not f) else ("female" if (f and not m) else "other")

        gen_i = None
        if gen is not None:
            gs = str(gen).strip()
            if gs.isdigit():
                gen_i = int(gs)

        raw_rows.append(
            {
                "row": r,
                "name": str(name).strip(),
                "gi": first_nonempty(gi),
                "spouse": str(spouse).strip() if has_spouse else "",
                "spouse_gi": first_nonempty(spouse_gi),
                "gen": gen_i,
                "gender": gender,
                "parent_cell": str(parent_cell).strip() if parent_cell else "",
                "owner_row": r,
            }
        )
        current_owner_row = r
    elif has_spouse:
        # Dòng phối ngẫu PHỤ: cột A trống nhưng cột C có tên
        # (vd Nguyễn Thị Loan - vợ hai của Hồ Tạo) -> bổ sung cho người ở dòng trước.
        if current_owner_row is not None:
            raw_rows.append(
                {
                    "row": r,
                    "name": "",
                    "gi": "",
                    "spouse": str(spouse).strip(),
                    "spouse_gi": first_nonempty(spouse_gi),
                    "gen": None,
                    "gender": "other",
                    "parent_cell": "",
                    "owner_row": current_owner_row,
                }
            )
    # nếu không tên và không phối ngẫu -> dòng cách, bỏ qua

# ---------------------------------------------------------------------------
# 0) Lan truyền cột H từ "hàng chốt" sang anh em cùng nhóm
# ---------------------------------------------------------------------------
# Cột H (Cha/Mẹ) chỉ điền ở hàng chốt mỗi nhóm. Các anh em ruột liền kề
# phía sau, CÙNG đời (cột E) và không có cột H riêng, sẽ kế thừa cha/mẹ của
# hàng chốt. Dừng khi gặp: dòng cách (gap theo số dòng thực), hàng chốt mới
# (có cột H riêng), hoặc đổi đời (cột E khác).

_named_rows = [rr for rr in raw_rows if rr["name"]]
_named_rows.sort(key=lambda x: x["row"])

# Dòng phối ngẫu phụ (cột A trống, cột C có tên - vd dòng 56 "Thị Lan") thuộc về
# người ở dòng trên, KHÔNG phải dòng cách ngăn nhóm.
_cont_rows = {rr["row"] for rr in raw_rows if not rr["name"]}

_i = 0
while _i < len(_named_rows):
    _anchor = _named_rows[_i]
    # Hàng chốt có thể bỏ trống cột E (vd dòng 159); anh em kế thừa vẫn hợp lệ
    # miễn cột E của họ giống hàng chốt (cùng None cũng tính là giống).
    if _anchor["parent_cell"]:
        _pc = _anchor["parent_cell"]
        _g = _anchor["gen"]
        _j = _i + 1
        while _j < len(_named_rows):
            _nxt = _named_rows[_j]
            # liền kề theo số dòng thực trong sheet (bắt được dòng trống ngăn nhóm);
            # bỏ qua các dòng phối ngẫu phụ nằm xen giữa
            if any(
                x not in _cont_rows
                for x in range(_named_rows[_j - 1]["row"] + 1, _nxt["row"])
            ):
                break
            if _nxt["parent_cell"]:
                break
            if _nxt["gen"] != _g:
                break
            _nxt["parent_cell"] = _pc
            _j += 1
        _i = _j
    else:
        _i += 1

# ---------------------------------------------------------------------------
# Mô hình person (chung cho cả người trong họ & dâu/rể)
# ---------------------------------------------------------------------------


class Person:
    __slots__ = (
        "id", "full_name", "gender", "gi", "gen_recorded",
        "gen_derived", "gen_final", "is_in_law", "variants",
        "canons", "note", "row",
    )

    def __init__(self, full_name, gender, gi, gen_recorded, is_in_law, row):
        self.id = str(uuid.uuid4())
        self.full_name = full_name
        self.gender = gender
        self.gi = gi
        self.gen_recorded = gen_recorded
        self.gen_derived = None
        self.gen_final = gen_recorded
        self.is_in_law = is_in_law
        self.row = row
        self.variants = []
        self.canons = set()
        self.note = None
        self._index_all()

    def _index_all(self):
        for n in [self.full_name] + self.variants:
            self.canons.add(canonical(n))

    def lunar(self):
        """Trả (day, month) từ chuỗi 'd.m' hoặc (None, None)."""
        if not self.gi:
            return None, None
        m = re.match(r"^(\d{1,2})[.](\d{1,2})$", self.gi.strip())
        if not m:
            return None, None
        return int(m.group(1)), int(m.group(2))


# ---------------------------------------------------------------------------
# 1) Người gốc trong họ (mỗi dòng dữ liệu = 1 người)
# ---------------------------------------------------------------------------

primaries: list[Person] = []
row_to_primary: dict[int, Person] = {}
for rr in raw_rows:
    if not rr["name"]:
        continue  # dòng phối ngẫu phụ (xử lý ở bước 2)
    variants = split_variants(rr["name"])
    primary = variants[0] if variants else rr["name"]
    p = Person(
        full_name=primary,
        gender=rr["gender"],
        gi=rr["gi"],
        gen_recorded=rr["gen"],
        is_in_law=False,
        row=rr["row"],
    )
    p.variants = variants[1:]
    p._index_all()
    p.note = f"Sheet1 dòng {rr['row']}"
    primaries.append(p)
    row_to_primary[rr["row"]] = p

index: dict[str, list[Person]] = {}
for p in primaries:
    for c in p.canons:
        index.setdefault(c, []).append(p)


def lookup(token: str) -> list[Person]:
    c = canonical(token)
    return index.get(c, [])


# Người họ Hồ (họ "ho" hay "hồ") và họ vợ-chồng dâu/rể khác nhau -> so đuôi tên thôi.
_HO = {"ho", "hồ"}


def resolve_parent(tok: str, child_row_gen, child_person: Person):
    """Tra cha/mẹ từ token cột H.

    Trả (person|None, ambiguous: bool, candidates: list[str]).
    Thứ tự: trùng canonical → trùng đuôi tên (đời = đời con - 1) → canonical kết thúc bằng token
    (cho trường hợp ghi "Thị Tư" = vợ "Phạm Thị Tư").
    """
    c = canonical(tok)
    if not c:
        return None, False, []

    def _resolve(cands: list[Person]):
        uniq: list[Person] = []
        seen_ids: set[str] = set()
        for x in cands:
            if x.id not in seen_ids:
                seen_ids.add(x.id)
                uniq.append(x)
        if len(uniq) == 1:
            return uniq[0], False, []
        p = _pick_by_gen(uniq, child_person, child_row_gen)
        if p:
            return p, False, []
        return None, True, [h.full_name + f"(đời {h.gen_recorded})" for h in uniq]

    # 1) trùng canonical - chỉ nhận khi đời khớp (cha phải ở đời con - 1).
    # Cột H ghi tên rút gọn ("Hồ Mân" = "Hồ Công Mân" đời 5) nên một canonical
    # trùng khít với người khác đời (Hồ Mân đời 6) là khớp giả -> đi tiếp bước 2.
    exact = index.get(c, [])
    if exact:
        g = child_person.gen_recorded if child_person.gen_recorded is not None else child_row_gen
        if g is None or any(
            (h.gen_recorded if h.gen_recorded is not None else h.gen_final) == g - 1
            for h in exact
        ):
            return _resolve(exact)

    c_tok = c.split()
    last = c_tok[-1]
    first = c_tok[0]

    # 2) trùng đuôi tên (token cuối + họ)
    cands: list[Person] = []
    seen: set[str] = set()
    for plist in index.values():
        for pp in plist:
            if pp.id in seen or pp.id == child_person.id:
                continue
            for cn in pp.canons:
                ct = cn.split()
                if ct and ct[-1] == last and ct[0] == first:
                    cands.append(pp)
                    seen.add(pp.id)
                    break
    if cands:
        res, amb, c2 = _resolve(cands)
        if res is not None or amb:
            return res, amb, c2

    # 3) canonical kết thúc bằng token (bắt tên vợ "Thị X" của chồng)
    if len(c) > 2:
        cands2 = [
            pp
            for plist in index.values()
            for pp in plist
            if pp.id != child_person.id
            and any(cn.endswith(c) for cn in pp.canons)
        ]
        if cands2:
            return _resolve(cands2)
    return None, False, []


def _pick_by_gen(cands, child_person, child_row_gen):
    g = child_row_gen if child_person.gen_recorded is None else child_person.gen_recorded
    if g is None:
        return None
    target = g - 1
    with_hint = [h for h in cands
                 if (h.gen_recorded if h.gen_recorded is not None else h.gen_final) == target]
    return with_hint[0] if len(with_hint) == 1 else None


# ---------------------------------------------------------------------------
# 2) Phối ngẫu (cột C): nối hôn nhân hoặc tạo person dâu/rể
# ---------------------------------------------------------------------------

inlaws: list[Person] = []
marriage_edges: list[tuple[Person, Person]] = []


def spouse_gender_hint(person: Person) -> str:
    if person.gender == "male":
        return "female"
    if person.gender == "female":
        return "male"
    return "other"


def ensure_spouse_person(token: str, gender_hint: str, gi: str, row: int) -> Person:
    # Chỉ match theo FULL NAME chính xác, không match variants/other_names.
    # "Thị Lan" ở nhiều dòng có thể là những người khác nhau; match qua
    # variants (vd "Dương Thị Soan / Thị Lan") sẽ merge nhầm người.
    variants = split_variants(token)
    primary_token = variants[0] if variants else token

    # Token KHÔNG CÓ HỌ (chỉ "Thị X" hoặc một từ) rất dễ trùng — luôn tạo
    # person mới, không match vào người đã có họ khác.
    tok_parts = primary_token.split()
    first_canon = canonical(tok_parts[0]) if tok_parts else ""
    has_family_name = len(tok_parts) >= 2 and first_canon not in ("thi",)

    if has_family_name:
        exact = lookup(primary_token)
        if exact:
            for cand in exact:
                if cand.full_name == primary_token:
                    return cand
            return exact[0]

    p = Person(
        full_name=primary_token,
        gender=gender_hint,
        gi=gi,
        gen_recorded=None,
        is_in_law=True,
        row=row,
    )
    p.variants = variants[1:]
    p._index_all()
    p.note = f"Phối ngẫu (cột C) - Sheet1 dòng {row}"
    inlaws.append(p)
    for c in p.canons:
        index.setdefault(c, []).append(p)
    return p


for rr in raw_rows:
    if not rr["spouse"]:
        continue
    # người sở hữu: dòng có tên chính mình, hoặc dòng phối ngẫu phụ (owner_row)
    owner_row = rr["row"] if rr["name"] else rr["owner_row"]
    p = row_to_primary.get(owner_row)
    if p is None:
        continue
    token = rr["spouse"]
    sp = ensure_spouse_person(token, spouse_gender_hint(p), rr["spouse_gi"], rr["row"])
    pair = tuple(sorted([p.id, sp.id]))
    if not any(sorted([a.id, b.id]) == list(pair) for a, b in marriage_edges):
        marriage_edges.append((p, sp))

# ---------------------------------------------------------------------------
# 3) Cha/Mẹ (cột H): nối biological_child
# ---------------------------------------------------------------------------

parent_edges: list[tuple[Person, Person]] = []  # (parent, child)
ambiguous_parents: list[tuple[int, str, list[str]]] = []

# Map spouse FULL NAME → person, và person → spouses (dùng để phân biệt cha
# trùng tên). Chỉ map theo full_name, KHÔNG theo variants/other_names để
# tránh "Thị Lan" bị nhầm thành "Dương Thị Soan / Thị Lan".
spouse_person_by_canon: dict[str, Person] = {}
person_spouses: dict[str, list[Person]] = {}
for a, b in marriage_edges:
    for x, y in ((a, b), (b, a)):
        person_spouses.setdefault(x.id, []).append(y)
        spouse_person_by_canon.setdefault(canonical(y.full_name), []).append(y)


def _spouse_of(p: Person) -> list[Person]:
    return person_spouses.get(p.id, [])


for rr in raw_rows:
    if not rr["name"] or not rr["parent_cell"]:
        continue
    p = row_to_primary[rr["row"]]
    tokens = [t.strip() for t in rr["parent_cell"].splitlines() if t.strip()]
    # Cột H ghi "Cha\nMẹ". Chỉ nối con với NGƯỜI ĐẦU TIÊN (cha — dòng họ),
    # mẹ/dâu đã được nối qua quan hệ marriage ở bước 2. Nối cả hai sẽ nhân
    # đôi cây và biến mẹ thành một gốc nhánh giả.
    if not tokens:
        continue
    tok = tokens[0]
    if canonical(tok) in p.canons:
        continue  # bỏ tự nối
    parent, ambiguous, cands = resolve_parent(tok, rr["gen"], p)

    # Khi cha trùng tên (ambiguous), dùng TÊN VỢ (token thứ 2 cột H) để chọn
    # cha có vợ khớp. VD "Hồ Hạnh\nThị Nhung" → cha = chồng của "Mai Thị Nhung".
    if (parent is None and ambiguous) and len(tokens) > 1:
        wife_tok = tokens[1]
        wife_hits = []
        for wc in spouse_person_by_canon:
            if canonical(wc).endswith(canonical(wife_tok)) or canonical(wife_tok).endswith(canonical(wc)):
                wife_hits.extend(spouse_person_by_canon[wc])
        # wife_hits là các person dâu/rể; tìm chồng (người trong họ) của họ
        husband_cands: list[Person] = []
        for w in wife_hits:
            for h in person_spouses.get(w.id, []):
                if not h.is_in_law:
                    husband_cands.append(h)
        if len(husband_cands) == 1:
            parent = husband_cands[0]
            ambiguous = False
            cands = []

    if parent is not None and parent.id != p.id:
        parent_edges.append((parent, p))
    elif ambiguous or parent is None:
        ambiguous_parents.append((rr["row"], tok, cands))


# ---------------------------------------------------------------------------
# 3b) Nối trục tổ chính (đời liền kề khi cột H trống)
# ---------------------------------------------------------------------------
# Các tổ đời đầu (vd Hồ Khang đời 1, Hồ Tạo đời 2) không có cột H. Nếu một
# người trong họ có đời ghi rõ (cột E) và ở đời NGAY TRƯỚC có DUY NHẤT một
# người trong họ, nối họ làm cha. Chỉ áp dụng cho người CHƯA có cha và
# chỉ trên trục chính (số người ở đời liền trước == 1) để tránh đoán bậy.

already_child_ids = {c.id for _p, c in parent_edges}

def _gen_of(p: Person):
    return p.gen_recorded if p.gen_recorded is not None else p.gen_final

primaries_by_gen: dict[int, list[Person]] = {}
for p in primaries:
    g = _gen_of(p)
    if g is not None:
        primaries_by_gen.setdefault(g, []).append(p)

for p in primaries:
    if p.id in already_child_ids:
        continue
    g = _gen_of(p)
    if g is None or g <= 0:
        continue
    prev = primaries_by_gen.get(g - 1, [])
    if len(prev) == 1 and prev[0].id != p.id:
        parent_edges.append((prev[0], p))


# ---------------------------------------------------------------------------
# 4) LOGIC CÁC ĐỜI: suy đời từ cha -> con (BFS), đối soát với cột E
# ---------------------------------------------------------------------------

all_persons = primaries + inlaws

# khởi tạo gen_parents (suy từ cha) = bản ghi cột E; gen_final = gen_parents
for p in all_persons:
    p.gen_derived = p.gen_recorded
    p.gen_final = p.gen_recorded

# đời dâu/rể (is_in_law) = đời của phối ngẫu
for a, b in marriage_edges:
    for x, y in ((a, b), (b, a)):
        if x.gen_final is None and y.gen_final is not None:
            x.gen_final = y.gen_final
            x.gen_derived = y.gen_final

# lan truyền cha -> con tới khi ổn định (đời con = đời cha + 1)
children_of: dict[str, list[Person]] = {}
for parent, child in parent_edges:
    children_of.setdefault(parent.id, []).append(child)

changed = True
while changed:
    changed = False
    for parent, child in parent_edges:
        pg = parent.gen_final if parent.gen_final is not None else parent.gen_derived
        if pg is None:
            continue
        # chỉ ghi khi chưa có (ưu tiên bản ghi cột E và đời dâu/rể đã có)
        if child.gen_final is None:
            child.gen_final = pg + 1
            child.gen_derived = pg + 1
            changed = True

# root = đời nhỏ nhất có bản ghi
valid = [q for q in all_persons if q.gen_final is not None]
root = min(valid, key=lambda q: q.gen_final) if valid else primaries[0]

# đời suy THUẦN từ cha (dùng để đối soát vs cột E, kể cả khi có bản ghi)
inferred: dict[str, int] = {}
for _ in range(len(all_persons)):
    changed_i = False
    for parent, child in parent_edges:
        pg = parent.gen_recorded if parent.gen_recorded is not None else inferred.get(parent.id)
        if pg is not None:
            val = pg + 1
            if inferred.get(child.id) != val:
                inferred[child.id] = val
                changed_i = True
    if not changed_i:
        break


# ---------------------------------------------------------------------------
# 5) Xuất JSON (format import v2)
# ---------------------------------------------------------------------------

persons_out = []
for p in all_persons:
    d, m = p.lunar()
    persons_out.append(
        {
            "id": p.id,
            "full_name": p.full_name,
            "gender": p.gender,
            "birth_year": None,
            "birth_month": None,
            "birth_day": None,
            "death_year": None,
            "death_month": None,
            "death_day": None,
            "death_lunar_year": None,
            "death_lunar_month": m,
            "death_lunar_day": d,
            "is_deceased": True,
            "is_in_law": p.is_in_law,
            "birth_order": None,
            "generation": p.gen_final,
            "other_names": ", ".join(p.variants) if p.variants else None,
            "avatar_url": None,
            "note": p.note,
        }
    )

rels_out = []
seen_marriage = set()
for a, b in marriage_edges:
    key = frozenset([a.id, b.id])
    if key in seen_marriage:
        continue
    seen_marriage.add(key)
    rels_out.append({"type": "marriage", "person_a": a.id, "person_b": b.id})

seen_child = set()
for parent, child in parent_edges:
    key = (parent.id, child.id)
    if key in seen_child:
        continue
    seen_child.add(key)
    rels_out.append({"type": "biological_child", "person_a": parent.id, "person_b": child.id})

payload = {
    "version": 2,
    "timestamp": "2026-08-13T00:00:00.000Z",
    "persons": persons_out,
    "relationships": rels_out,
}
OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# 6) Báo cáo đối soát logic các đời
# ---------------------------------------------------------------------------

generation_conflicts = []
generation_filled = []
generation_missing = []

for p in all_persons:
    rec = p.gen_recorded
    der = inferred.get(p.id)
    if rec is not None and der is not None and rec != der:
        generation_conflicts.append((p, rec, der))
    elif rec is None and der is not None:
        generation_filled.append((p, der))
    elif rec is None and p.gen_final is None:
        generation_missing.append(p)


def rowfmt(p: Person) -> str:
    return f"`{p.full_name}` (dòng {p.row})"


lines = []
lines.append("# Đối soát logic các đời - từ `Gia phả.xlsx`\n")
lines.append(f"- Người trong họ: **{len(primaries)}** · Dâu/rể tự sinh: **{len(inlaws)}** · Tổng person: **{len(all_persons)}**\n")
lines.append(f"- Cạnh cha→con: **{len(parent_edges)}** · Cạnh hôn nhân: **{len(marriage_edges)}**\n")
lines.append(f"- Root: **{root.full_name}** (đời {root.gen_recorded})\n")
lines.append("- File JSON: `plans/260619-1933-digitize-genealogy/giapha-from-xlsx.json`\n")

lines.append("\n## 1. Cây theo từng nhánh (cha → con)\n")
rendered = set()
# các nhánh: bắt đầu từ người làm cha/mẹ nhưng không phải con ai cả
child_ids = {c.id for _p, c in parent_edges}
all_parent_ids = {p.id for p, _c in parent_edges}
top_nodes = [p for p in all_persons if p.id in all_parent_ids and p.id not in child_ids]
if not top_nodes:
    top_nodes = [root]
for top in sorted(top_nodes, key=lambda q: (q.gen_final if q.gen_final is not None else 99, q.full_name)):
    lines.append(f"\n**Nhánh {top.full_name}**\n")
    stack = [(top, 0)]
    while stack:
        cur, depth = stack.pop()
        if cur.id in rendered:
            continue
        rendered.add(cur.id)
        gen_txt = cur.gen_final if cur.gen_final is not None else "?"
        line = ("  " * depth) + f"{gen_txt} {cur.full_name}"
        if cur.is_in_law:
            line += " *(dâu)*"
        if cur.gi:
            line += f" [giỗ {cur.gi}]"
        lines.append(line)
        kids = children_of.get(cur.id, [])
        for k in reversed(kids):
            if k.id not in rendered:
                stack.append((k, depth + 1))

if generation_conflicts:
    lines.append("\n## 2. ⚠️ Xung đột đời (cột E ≠ đời suy từ cha)\n")
    for p, rec, der in generation_conflicts:
        lines.append(f"- {rowfmt(p)}: bản ghi **{rec}** nhưng cha-mẹ cho **{der}**")
else:
    lines.append("\n## 2. ✅ Không có xung đột đời nào.\n")

if generation_filled:
    lines.append("\n## 3. Điền đời còn thiếu (từ cha/mẹ)\n")
    for p, der in generation_filled:
        lines.append(f"- {rowfmt(p)} → đời **{der}**")
else:
    lines.append("\n## 3. Không cần điền đời nào.\n")

if generation_missing:
    lines.append("\n## 4. ❓ Vẫn còn người thiếu đời (chưa tự suy được)\n")
    for p in generation_missing:
        lines.append(f"- {rowfmt(p)}")
else:
    lines.append("\n## 4. ✅ Không còn người thiếu đời.\n")

if ambiguous_parents:
    lines.append("\n## 5. ⚠️ Cha/mẹ mờ (không nối được chắc chắn)\n")
    for row, tok, cands in ambiguous_parents:
        detail = ", ".join(cands) if cands else "không tìm thấy trong bảng"
        lines.append(f"- dòng {row}: `{tok}` → {detail}")
else:
    lines.append("\n## 5. ✅ Tất cả cha/mẹ đều nối rõ ràng.\n")

name_counts: dict[str, list[Person]] = {}
for p in primaries:
    name_counts.setdefault(p.full_name, []).append(p)
dups = {n: v for n, v in name_counts.items() if len(v) > 1}
if dups:
    lines.append("\n## 6. ℹ️ Trùng tên (nhiều người cùng tên chính)\n")
    for n, v in dups.items():
        lines.append(f"- **{n}** ({len(v)} người): " + ", ".join(rowfmt(p) for p in v))
else:
    lines.append("\n## 6. ✅ Không có tên trùng.\n")

lines.append("\n## 7. Quy ước\n")
lines.append("- Cột B **Ngày sinh** (dạng `d.m`) coi là **ngày giỗ âm lịch** → ghi `death_lunar_day/month`, đồng nhất bản `giapha-backbone-draft.json` trước đó.")
lines.append("- Mọi person đặt `is_deceased = true` (đồng bộ quy ước cũ); người còn sống có thể sửa trong app.")
lines.append("- Người cột C (vợ/chồng) là **dâu/rể** → `is_in_law = true`; người cột A là người trong họ.")
lines.append("- Tên dạng `A/B` → `full_name = A`, `other_names = B`.")
lines.append("- Đời (generation) ưu tiên cột E `Thế hệ thứ`; nếu trống, suy từ `đời cha + 1` (BFS từ root).")

OUT_REPORT.write_text("\n".join(lines), encoding="utf-8")

print("Done.")
print(f"  persons: {len(persons_out)}  relationships: {len(rels_out)}")
print(f"  conflicts: {len(generation_conflicts)}  filled: {len(generation_filled)}"
      f"  missing: {len(generation_missing)}  ambiguous parents: {len(ambiguous_parents)}")
print(f"  JSON: {OUT_JSON}")
print(f"  Report: {OUT_REPORT}")

